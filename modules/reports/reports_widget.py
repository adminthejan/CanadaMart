"""Sales reports and analytics module."""
from datetime import datetime, timedelta
from typing import List, Dict

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QDateEdit,
    QComboBox, QFrame, QGridLayout, QSizePolicy, QFileDialog,
    QScrollArea, QMessageBox, QTabWidget,
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QColor, QBrush, QFont

try:
    import matplotlib
    matplotlib.use("QtAgg")
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False


# ═══════════════════════════════════════════════════════════════════════════
#  Stat card
# ═══════════════════════════════════════════════════════════════════════════
def make_stat_card(title: str, value: str, subtitle: str = "", color: str = "#f1f5f9") -> QFrame:
    card = QFrame()
    card.setObjectName("StatCard")
    layout = QVBoxLayout(card)
    layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

    val = QLabel(value)
    val.setObjectName("StatValue")
    val.setStyleSheet(f"color:{color}; font-size:22px; font-weight:700;")
    val.setAlignment(Qt.AlignmentFlag.AlignCenter)
    layout.addWidget(val)

    lbl = QLabel(title)
    lbl.setObjectName("StatLabel")
    lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
    layout.addWidget(lbl)

    if subtitle:
        sub = QLabel(subtitle)
        sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sub.setStyleSheet("color:#64748b; font-size:10px;")
        layout.addWidget(sub)

    return card


# ═══════════════════════════════════════════════════════════════════════════
#  Charts widget
# ═══════════════════════════════════════════════════════════════════════════
class SalesChart(QWidget):
    """Matplotlib-powered line chart for daily sales."""

    def __init__(self):
        super().__init__()
        if not HAS_MATPLOTLIB:
            lbl = QLabel("Install matplotlib for charts:\npip install matplotlib")
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl.setStyleSheet("color:#64748b;")
            QVBoxLayout(self).addWidget(lbl)
            return

        self._fig = Figure(figsize=(8, 3), facecolor="#1e293b")
        self._ax = self._fig.add_subplot(111)
        self._ax.set_facecolor("#0f172a")
        self._canvas = FigureCanvas(self._fig)
        QVBoxLayout(self).addWidget(self._canvas)

    def plot_daily(self, daily_data: List[dict], currency_sym: str = "$"):
        if not HAS_MATPLOTLIB or not daily_data:
            return
        dates = [d["date"] for d in daily_data]
        revenues = [float(d.get("revenue", 0)) for d in daily_data]
        txns = [int(d.get("transactions", 0)) for d in daily_data]

        self._ax.clear()
        self._ax.set_facecolor("#0f172a")
        self._fig.patch.set_facecolor("#1e293b")

        self._ax.plot(dates, revenues, color="#3b82f6", linewidth=2,
                      marker="o", markersize=4, label="Revenue")
        self._ax.fill_between(dates, revenues, alpha=0.1, color="#3b82f6")

        ax2 = self._ax.twinx()
        ax2.bar(dates, txns, alpha=0.3, color="#8b5cf6", label="Transactions")
        ax2.set_ylabel("Transactions", color="#8b5cf6", fontsize=9)
        ax2.tick_params(colors="#64748b")

        self._ax.set_xlabel("")
        self._ax.set_ylabel(f"Revenue ({currency_sym})", color="#94a3b8", fontsize=9)
        self._ax.tick_params(axis="x", rotation=45, colors="#64748b", labelsize=7)
        self._ax.tick_params(axis="y", colors="#64748b")
        self._ax.spines["bottom"].set_color("#334155")
        self._ax.spines["left"].set_color("#334155")
        self._ax.spines["top"].set_visible(False)
        self._ax.spines["right"].set_visible(False)
        ax2.spines["top"].set_visible(False)
        ax2.spines["right"].set_color("#334155")

        if len(dates) > 7:
            step = max(1, len(dates) // 7)
            self._ax.set_xticks(self._ax.get_xticks()[::step])

        self._fig.tight_layout()
        self._canvas.draw()


# ═══════════════════════════════════════════════════════════════════════════
#  Main Reports Widget
# ═══════════════════════════════════════════════════════════════════════════
class ReportsWidget(QWidget):
    def __init__(self, config, db):
        super().__init__()
        self.config = config
        self.db = db
        self._build_ui()
        self._load_today()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        # ── Toolbar ─────────────────────────────────────────────────────
        toolbar = QHBoxLayout()
        title = QLabel("📊  Sales Reports")
        title.setStyleSheet("font-size:20px; font-weight:700; color:#f1f5f9;")
        toolbar.addWidget(title)
        toolbar.addStretch()

        # Quick range buttons
        for label, days in [("Today", 0), ("7 Days", 7), ("30 Days", 30), ("This Month", -1)]:
            btn = QPushButton(label)
            btn.setObjectName("SecondaryBtn")
            btn.clicked.connect(lambda _, d=days: self._quick_range(d))
            toolbar.addWidget(btn)

        toolbar.addWidget(QLabel("  From:"))
        self._date_from = QDateEdit()
        self._date_from.setCalendarPopup(True)
        self._date_from.setDate(QDate.currentDate())
        toolbar.addWidget(self._date_from)

        toolbar.addWidget(QLabel("To:"))
        self._date_to = QDateEdit()
        self._date_to.setCalendarPopup(True)
        self._date_to.setDate(QDate.currentDate())
        toolbar.addWidget(self._date_to)

        load_btn = QPushButton("🔍 Load")
        load_btn.setObjectName("PrimaryBtn")
        load_btn.clicked.connect(self._load_report)
        toolbar.addWidget(load_btn)

        export_btn = QPushButton("📤 Export CSV")
        export_btn.clicked.connect(self._export)
        toolbar.addWidget(export_btn)

        pdf_btn = QPushButton("📄 Download PDF")
        pdf_btn.setObjectName("SecondaryBtn")
        pdf_btn.clicked.connect(self._export_pdf)
        toolbar.addWidget(pdf_btn)

        layout.addLayout(toolbar)

        # ── Summary cards ────────────────────────────────────────────────
        self._cards_layout = QHBoxLayout()
        self._card_revenue = make_stat_card("Total Revenue", "$0.00", color="#4ade80")
        self._card_txns = make_stat_card("Transactions", "0", color="#60a5fa")
        self._card_avg = make_stat_card("Avg. Sale", "$0.00", color="#a78bfa")
        self._card_tax = make_stat_card("Tax Collected", "$0.00", color="#fb923c")
        self._card_disc = make_stat_card("Discounts Given", "$0.00", color="#f87171")
        for c in [self._card_revenue, self._card_txns, self._card_avg, self._card_tax, self._card_disc]:
            self._cards_layout.addWidget(c)
        layout.addLayout(self._cards_layout)

        # ── Tabs: Chart | Sales List | Top Products | Payment Breakdown ──
        tabs = QTabWidget()
        layout.addWidget(tabs, 1)

        # Chart tab
        self._chart = SalesChart()
        chart_scroll = QScrollArea()
        chart_scroll.setWidget(self._chart)
        chart_scroll.setWidgetResizable(True)
        tabs.addTab(chart_scroll, "📈 Revenue Chart")

        # Sales list tab
        self._sales_table = QTableWidget(0, 7)
        self._sales_table.setHorizontalHeaderLabels([
            "Invoice", "Date", "Customer", "Items", "Subtotal",
            "Tax", "Total", "Method"
        ])
        hh = self._sales_table.horizontalHeader()
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        for i in [0, 1, 3, 4, 5, 6, 7]:
            hh.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        self._sales_table.verticalHeader().hide()
        self._sales_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._sales_table.setAlternatingRowColors(True)
        self._sales_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        tabs.addTab(self._sales_table, "🧾 Sales List")

        # Top products tab
        self._top_table = QTableWidget(0, 4)
        self._top_table.setHorizontalHeaderLabels(["Product", "SKU", "Units Sold", "Revenue"])
        hh2 = self._top_table.horizontalHeader()
        hh2.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self._top_table.verticalHeader().hide()
        self._top_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._top_table.setAlternatingRowColors(True)
        tabs.addTab(self._top_table, "🏆 Top Products")

        # Payment breakdown tab
        self._pay_table = QTableWidget(0, 3)
        self._pay_table.setHorizontalHeaderLabels(["Payment Method", "Count", "Total"])
        hh3 = self._pay_table.horizontalHeader()
        hh3.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self._pay_table.verticalHeader().hide()
        self._pay_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        tabs.addTab(self._pay_table, "💳 Payment Methods")

    # ------------------------------------------------------------------ #
    #  Data loading                                                        #
    # ------------------------------------------------------------------ #
    def refresh(self):
        """Called whenever the Reports tab becomes visible – reload current range."""
        self._load_report()

    def _load_today(self):
        today = QDate.currentDate()
        self._date_from.setDate(today)
        self._date_to.setDate(today)
        self._load_report()

    def _quick_range(self, days: int):
        today = QDate.currentDate()
        if days == 0:
            self._date_from.setDate(today)
        elif days == -1:
            self._date_from.setDate(QDate(today.year(), today.month(), 1))
        else:
            self._date_from.setDate(today.addDays(-days))
        self._date_to.setDate(today)
        self._load_report()

    def _load_report(self):
        date_from = self._date_from.date().toString("yyyy-MM-dd")
        date_to = self._date_to.date().toString("yyyy-MM-dd")
        sym = self.config.get("currency_symbol", "$")

        # Summary
        summary = self.db.get_sales_summary(date_from, date_to)
        revenue = float(summary.get("total_revenue", 0) or 0)
        txns = int(summary.get("total_transactions", 0) or 0)
        avg = float(summary.get("avg_transaction", 0) or 0)
        tax = float(summary.get("total_tax", 0) or 0)
        disc = float(summary.get("total_discounts", 0) or 0)

        # Update stat cards
        def _update_card(card: QFrame, value: str):
            for lbl in card.findChildren(QLabel):
                if lbl.objectName() == "StatValue":
                    lbl.setText(value)
                    break

        _update_card(self._card_revenue, f"{sym}{revenue:.2f}")
        _update_card(self._card_txns, str(txns))
        _update_card(self._card_avg, f"{sym}{avg:.2f}")
        _update_card(self._card_tax, f"{sym}{tax:.2f}")
        _update_card(self._card_disc, f"{sym}{disc:.2f}")

        # Chart
        daily = self.db.get_daily_sales(date_from, date_to)
        self._chart.plot_daily(daily, sym)

        # Sales list
        sales = self.db.get_sales(date_from, date_to)
        self._sales_table.setRowCount(len(sales))
        for row, sale in enumerate(sales):
            items = self.db.get_sale_items(sale["id"])
            self._sales_table.setItem(row, 0, QTableWidgetItem(sale.get("invoice_number", "")))
            self._sales_table.setItem(row, 1, QTableWidgetItem(sale.get("created_at", "")[:16]))
            self._sales_table.setItem(row, 2, QTableWidgetItem(sale.get("customer_name", "Walk-in")))
            cnt = QTableWidgetItem(str(len(items)))
            cnt.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self._sales_table.setItem(row, 3, cnt)
            self._sales_table.setItem(row, 4, QTableWidgetItem(f"{sym}{sale.get('subtotal',0):.2f}"))
            self._sales_table.setItem(row, 5, QTableWidgetItem(f"{sym}{sale.get('tax_amount',0):.2f}"))
            total_item = QTableWidgetItem(f"{sym}{sale.get('total',0):.2f}")
            total_item.setForeground(QBrush(QColor("#4ade80")))
            self._sales_table.setItem(row, 6, total_item)
            self._sales_table.setItem(row, 7, QTableWidgetItem(sale.get("payment_method", "").upper()))
        self._sales_table.resizeRowsToContents()

        # Top products
        top = self.db.get_top_products(date_from, date_to)
        self._top_table.setRowCount(len(top))
        for row, p in enumerate(top):
            self._top_table.setItem(row, 0, QTableWidgetItem(p.get("product_name", "")))
            self._top_table.setItem(row, 1, QTableWidgetItem(p.get("sku", "") or ""))
            units_item = QTableWidgetItem(str(p.get("units_sold", 0)))
            units_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self._top_table.setItem(row, 2, units_item)
            rev_item = QTableWidgetItem(f"{sym}{p.get('revenue',0):.2f}")
            rev_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            rev_item.setForeground(QBrush(QColor("#4ade80")))
            self._top_table.setItem(row, 3, rev_item)
        self._top_table.resizeRowsToContents()

        # Payment breakdown
        payments = self.db.get_payment_breakdown(date_from, date_to)
        self._pay_table.setRowCount(len(payments))
        for row, pay in enumerate(payments):
            self._pay_table.setItem(row, 0, QTableWidgetItem(pay.get("payment_method", "").upper()))
            cnt_item = QTableWidgetItem(str(pay.get("count", 0)))
            cnt_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self._pay_table.setItem(row, 1, cnt_item)
            total_item = QTableWidgetItem(f"{sym}{pay.get('total',0):.2f}")
            total_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self._pay_table.setItem(row, 2, total_item)

    # ------------------------------------------------------------------ #
    #  PDF Export (all reports)                                            #
    # ------------------------------------------------------------------ #
    def _export_pdf(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Report PDF", "sales_report.pdf",
            "PDF Files (*.pdf)"
        )
        if not path:
            return

        date_from = self._date_from.date().toString("yyyy-MM-dd")
        date_to   = self._date_to.date().toString("yyyy-MM-dd")
        sym       = self.config.get("currency_symbol", "$")

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import mm
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table,
                TableStyle, HRFlowable, KeepTogether,
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
        except ImportError:
            QMessageBox.warning(self, "Missing Package",
                                "Install reportlab:\npip install reportlab")
            return

        import os

        doc = SimpleDocTemplate(
            path, pagesize=A4,
            leftMargin=15*mm, rightMargin=15*mm,
            topMargin=15*mm, bottomMargin=15*mm,
        )

        styles = getSampleStyleSheet()
        h1  = ParagraphStyle("h1",  parent=styles["Heading1"],
                             fontSize=16, textColor=colors.HexColor("#1e3a5f"),
                             spaceAfter=4)
        h2  = ParagraphStyle("h2",  parent=styles["Heading2"],
                             fontSize=12, textColor=colors.HexColor("#1e3a5f"),
                             spaceBefore=10, spaceAfter=4)
        ctr = ParagraphStyle("ctr", parent=styles["Normal"],
                             alignment=TA_CENTER, fontSize=9,
                             textColor=colors.HexColor("#374151"))
        sml = ParagraphStyle("sml", parent=styles["Normal"],
                             fontSize=8, textColor=colors.HexColor("#374151"))
        biz_name   = self.config.get("business_name", "")
        biz_addr   = self.config.get("business_address", "")
        biz_phone  = self.config.get("business_phone", "")
        logo_path  = self.config.get("logo_path", "")
        show_logo  = self.config.get("receipt_show_logo", True)

        story = []

        # ── Header: logo + business info ─────────────────────────────── #
        if show_logo and logo_path and os.path.exists(logo_path):
            try:
                from reportlab.platypus import Image as RLImage
                logo = RLImage(logo_path, width=40*mm, height=20*mm,
                               kind="proportional")
                logo.hAlign = "CENTER"
                story.append(logo)
                story.append(Spacer(1, 3*mm))
            except Exception:
                pass

        story.append(Paragraph(biz_name or "Sales Report", h1))
        if biz_addr:
            story.append(Paragraph(biz_addr, ctr))
        if biz_phone:
            story.append(Paragraph(biz_phone, ctr))
        story.append(Paragraph(
            f"Report Period: <b>{date_from}</b> to <b>{date_to}</b>", ctr))
        story.append(HRFlowable(width="100%",
                                color=colors.HexColor("#1e3a5f"), thickness=1.5))
        story.append(Spacer(1, 5*mm))

        # ── Summary cards ────────────────────────────────────────────── #
        summary = self.db.get_sales_summary(date_from, date_to)
        revenue = float(summary.get("total_revenue", 0) or 0)
        txns    = int(summary.get("total_transactions", 0) or 0)
        avg     = float(summary.get("avg_transaction", 0) or 0)
        tax     = float(summary.get("total_tax", 0) or 0)
        disc    = float(summary.get("total_discounts", 0) or 0)

        story.append(Paragraph("Summary", h2))
        summary_data = [
            ["Metric", "Value"],
            ["Total Revenue",     f"{sym}{revenue:.2f}"],
            ["Transactions",      str(txns)],
            ["Average Sale",      f"{sym}{avg:.2f}"],
            ["Tax Collected",     f"{sym}{tax:.2f}"],
            ["Discounts Given",   f"{sym}{disc:.2f}"],
        ]
        col_w = [100*mm, 60*mm]
        sum_tbl = Table(summary_data, colWidths=col_w)
        sum_tbl.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 9),
            ("ALIGN",       (1, 0), (1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.HexColor("#f8fafc"), colors.HexColor("#e2e8f0")]),
            ("LINEBELOW",   (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",(0, 0), (-1, -1), 6),
            ("TOPPADDING",  (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 0),(-1, -1), 4),
        ]))
        story.append(sum_tbl)
        story.append(Spacer(1, 6*mm))

        # ── Sales list ───────────────────────────────────────────────── #
        sales = self.db.get_sales(date_from, date_to)
        story.append(Paragraph(f"Sales List ({len(sales)} transactions)", h2))

        if sales:
            tdata = [["Invoice", "Date", "Customer", "Subtotal",
                      "Tax", "Discount", "Total", "Method"]]
            for s in sales:
                tdata.append([
                    s.get("invoice_number", ""),
                    (s.get("created_at", "") or "")[:16],
                    s.get("customer_name", "Walk-in") or "Walk-in",
                    f"{sym}{s.get('subtotal', 0):.2f}",
                    f"{sym}{s.get('tax_amount', 0):.2f}",
                    f"{sym}{s.get('discount_amount', 0):.2f}",
                    f"{sym}{s.get('total', 0):.2f}",
                    (s.get("payment_method", "") or "").upper(),
                ])
            pw = A4[0] - 30*mm
            cws = [pw*0.13, pw*0.14, pw*0.18, pw*0.12,
                   pw*0.10, pw*0.10, pw*0.12, pw*0.11]
            stbl = Table(tdata, colWidths=cws, repeatRows=1)
            stbl.setStyle(TableStyle([
                ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
                ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
                ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",    (0, 0), (-1, -1), 7),
                ("ALIGN",       (3, 0), (-1, -1), "RIGHT"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.white, colors.HexColor("#f1f5f9")]),
                ("LINEBELOW",   (0, 0), (-1, -1), 0.2, colors.HexColor("#e2e8f0")),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING",(0, 0), (-1, -1), 4),
                ("TOPPADDING",  (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING",(0, 0),(-1, -1), 3),
            ]))
            story.append(stbl)
        else:
            story.append(Paragraph("No sales in this period.", sml))

        story.append(Spacer(1, 6*mm))

        # ── Top Products ─────────────────────────────────────────────── #
        top = self.db.get_top_products(date_from, date_to)
        story.append(Paragraph(f"Top Products ({len(top)} items)", h2))

        if top:
            tpdata = [["Product", "SKU", "Units Sold", "Revenue"]]
            for p in top:
                tpdata.append([
                    p.get("product_name", ""),
                    p.get("sku", "") or "",
                    str(p.get("units_sold", 0)),
                    f"{sym}{p.get('revenue', 0):.2f}",
                ])
            pw = A4[0] - 30*mm
            tpcws = [pw*0.50, pw*0.20, pw*0.15, pw*0.15]
            tptbl = Table(tpdata, colWidths=tpcws, repeatRows=1)
            tptbl.setStyle(TableStyle([
                ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
                ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
                ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",    (0, 0), (-1, -1), 8),
                ("ALIGN",       (2, 0), (-1, -1), "RIGHT"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.white, colors.HexColor("#f1f5f9")]),
                ("LINEBELOW",   (0, 0), (-1, -1), 0.2, colors.HexColor("#e2e8f0")),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING",(0, 0), (-1, -1), 5),
                ("TOPPADDING",  (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING",(0, 0),(-1, -1), 3),
            ]))
            story.append(tptbl)
        else:
            story.append(Paragraph("No product data in this period.", sml))

        story.append(Spacer(1, 6*mm))

        # ── Payment Breakdown ─────────────────────────────────────────── #
        payments = self.db.get_payment_breakdown(date_from, date_to)
        story.append(Paragraph("Payment Methods", h2))

        if payments:
            ppdata = [["Payment Method", "Transactions", "Total"]]
            for p in payments:
                ppdata.append([
                    (p.get("payment_method", "") or "").upper(),
                    str(p.get("count", 0)),
                    f"{sym}{p.get('total', 0):.2f}",
                ])
            pw = A4[0] - 30*mm
            ppcws = [pw*0.50, pw*0.25, pw*0.25]
            pptbl = Table(ppdata, colWidths=ppcws)
            pptbl.setStyle(TableStyle([
                ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
                ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
                ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",    (0, 0), (-1, -1), 9),
                ("ALIGN",       (1, 0), (-1, -1), "RIGHT"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.white, colors.HexColor("#f1f5f9")]),
                ("LINEBELOW",   (0, 0), (-1, -1), 0.2, colors.HexColor("#e2e8f0")),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING",(0, 0), (-1, -1), 5),
                ("TOPPADDING",  (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING",(0, 0),(-1, -1), 4),
            ]))
            story.append(pptbl)
        else:
            story.append(Paragraph("No payment data in this period.", sml))

        # ── Footer ───────────────────────────────────────────────────── #
        story.append(Spacer(1, 8*mm))
        story.append(HRFlowable(width="100%",
                                color=colors.HexColor("#cbd5e1"), thickness=0.5))
        from datetime import datetime as _dt
        story.append(Paragraph(
            f"Generated: {_dt.now().strftime('%d-%b-%Y %H:%M')}  |  "
            f"{biz_name}",
            ParagraphStyle("foot", parent=styles["Normal"],
                           fontSize=7, textColor=colors.HexColor("#94a3b8"),
                           alignment=TA_CENTER)
        ))

        try:
            doc.build(story)
            QMessageBox.information(self, "PDF Saved", f"Report saved to:\n{path}")
            import subprocess, sys
            if sys.platform == "win32":
                import os as _os; _os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", path])
            else:
                subprocess.Popen(["xdg-open", path])
        except Exception as e:
            QMessageBox.critical(self, "PDF Error", f"Could not create PDF:\n{e}")

    # ------------------------------------------------------------------ #
    #  CSV Export                                                          #
    # ------------------------------------------------------------------ #
    def _export(self):
        path, filt = QFileDialog.getSaveFileName(
            self, "Export Report", "sales_report.csv",
            "CSV Files (*.csv);;Excel Files (*.xlsx)"
        )
        if not path:
            return

        date_from = self._date_from.date().toString("yyyy-MM-dd")
        date_to = self._date_to.date().toString("yyyy-MM-dd")
        sales = self.db.get_sales(date_from, date_to)
        sym = self.config.get("currency_symbol", "$")

        if path.endswith(".xlsx"):
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Sales Report"
                headers = ["Invoice", "Date", "Customer", "Subtotal",
                           "Tax", "Discount", "Total", "Method", "Notes"]
                ws.append(headers)
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="1e293b")
                for sale in sales:
                    ws.append([
                        sale.get("invoice_number", ""),
                        sale.get("created_at", "")[:16],
                        sale.get("customer_name", ""),
                        sale.get("subtotal", 0),
                        sale.get("tax_amount", 0),
                        sale.get("discount_amount", 0),
                        sale.get("total", 0),
                        sale.get("payment_method", ""),
                        sale.get("notes", ""),
                    ])
                wb.save(path)
                QMessageBox.information(self, "Export Complete", f"Saved to {path}")
            except ImportError:
                QMessageBox.warning(self, "Missing Package",
                                    "Install openpyxl for Excel export:\npip install openpyxl")
        else:
            import csv
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Invoice", "Date", "Customer", "Subtotal",
                                  "Tax", "Discount", "Total", "Method", "Notes"])
                for sale in sales:
                    writer.writerow([
                        sale.get("invoice_number", ""),
                        sale.get("created_at", "")[:16],
                        sale.get("customer_name", ""),
                        sale.get("subtotal", 0),
                        sale.get("tax_amount", 0),
                        sale.get("discount_amount", 0),
                        sale.get("total", 0),
                        sale.get("payment_method", ""),
                        sale.get("notes", ""),
                    ])
            QMessageBox.information(self, "Export Complete", f"Saved to {path}")
